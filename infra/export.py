from __future__ import annotations
import io
import pandas as pd
from openpyxl.styles import numbers


def dataframe_a_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Conciliacion",
    formato_columnas_fecha: dict[str, str] | None = None
) -> bytes:
    """
    Exporta un DataFrame a Excel conservando los tipos fecha (no texto).
    Si se pasa `formato_columnas_fecha` con {nombre_columna: "DD/MM/YYYY"}, aplica number_format.
    """
    buff = io.BytesIO()
    with pd.ExcelWriter(buff, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        if formato_columnas_fecha:
            ws = writer.sheets[sheet_name]
            # Mapear nombres de columnas a letras
            headers = [c.value for c in ws[1]]
            for col_name, fmt in formato_columnas_fecha.items():
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for cell in ws[col_letter][1:]:
                        cell.number_format = fmt
    return buff.getvalue()
